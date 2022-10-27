import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pandas as pd

filename1 = 'data/sales_one.xlsx'
filename2 = 'data/sales_two.xlsx'

df_one = pd.read_excel(filename1, index_col=0)
df_two = pd.read_excel(filename2, index_col=0)

# merge das planilhas eliminando as linhas duplicadas
df_merge = df_one.merge(df_two, how='outer')
# print(df_merge)

# Criação da nova planilha mergeada sem dados duplicados
filename = 'data/sales_merge.xlsx'
df_merge.to_excel(filename, sheet_name="Sales merged")

# Verificação dos dados criados
df_new = pd.read_excel(filename, index_col=0)
# print(df_new)

# Escrevendo na planilha criada
workbook = openpyxl.load_workbook(filename)
worksheet = workbook['Sales merged']
last_index = str(len(df_merge)+2)
# print(last_index)

# Adicionando o título ao campo
worksheet['A'+last_index] = 'Total of Sales'
# Formatando a fonte e o campo
worksheet['A' +
          last_index].font = openpyxl.styles.Font(bold=True, color="00FFFFFF")
worksheet['A'+last_index].alignment = Alignment(horizontal="center",
                                                vertical="center")
worksheet['A'+last_index].fill = PatternFill("solid", fgColor="000000FF")

# Calculando o total e escrevendo na posição D o resultado da fdórmula do Excel
worksheet['D'+last_index] = '=SUMPRODUCT(C2:C{},D2:D{})'.format(
    str(int(last_index)-1), str(int(last_index)-1))

# Formatando a fonte do valor total incluindo o formato monetário
worksheet['D'+last_index].font = openpyxl.styles.Font(bold=True)
worksheet['D'+last_index].number_format = '#,##0.00$'
workbook.save(filename)
